from django.test import TestCase, Client
from django.urls import reverse
from .models import Robot
import json
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from openpyxl import load_workbook



class RobotCreateViewTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse('robot_create')
        self.valid_payload = {
            "model": "R2",
            "version": "D2",
            "created": "2022-12-31T23:59:59Z"
        }

    def test_create_robot_success(self):
        response = self.client.post(
            path=self.url,
            data=json.dumps(self.valid_payload),
            content_type='application/json'
        )
        
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Robot.objects.count(), 1)
        robot = Robot.objects.first()
        self.assertEqual(robot.model, "R2")
        self.assertEqual(robot.version, "D2")
        self.assertEqual(str(robot.created), "2022-12-31 23:59:59+00:00")

        response_data = response.json()
        self.assertEqual(response_data['model'], "R2")
        self.assertEqual(response_data['version'], "D2")
        self.assertEqual(response_data['created'], "2022-12-31T23:59:59Z")
        self.assertEqual(response_data['serial'], "R2-D2")

    def test_create_robot_invalid_data(self):
        invalid_payload = {
            "version": "D2",
            "created": "2022-12-31T23:59:59Z"
        }
        
        response = self.client.post(
            path=self.url,
            data=json.dumps(invalid_payload),
            content_type='application/json'
        )
        
        self.assertEqual(response.status_code, 400)
        response_data = response.json()
        self.assertIn('error', response_data)
        self.assertEqual(response_data['error'], "Поля model, version и created обязательны.")

class RobotReportViewTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.report_url = reverse('robot_report')
        now = timezone.now()

        # Создаем тестовые данные
        self.robot_data = [
            {"model": "MX", "version": "v1", "created": now},
            {"model": "MX", "version": "v1", "created": now},
            {"model": "MX", "version": "v1", "created": now},
            {"model": "MX", "version": "v2", "created": now - timedelta(days=2)},
            {"model": "MY", "version": "v1", "created": now - timedelta(days=4)},
            {"model": "MZ", "version": "v1", "created": now - timedelta(days=6)},
            {"model": "MZ", "version": "v2", "created": now},
        ]

        for data in self.robot_data:
            Robot.objects.create(**data)

    def test_generate_report_success(self):
        response = self.client.get(self.report_url)
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        file_content = response.content  # Читаем весь контент из HttpResponse
        workbook = load_workbook(filename=BytesIO(file_content))
        sheet_names = workbook.sheetnames

        # Проверяем, что в отчете есть страницы с именами моделей
        self.assertIn("MX", sheet_names)
        self.assertIn("MY", sheet_names)
        self.assertIn("MZ", sheet_names)

        # Проверяем содержимое страницы MX
        sheet = workbook["MX"]
        self.assertEqual(sheet.cell(row=2, column=1).value, "MX")
        self.assertIn(sheet.cell(row=2, column=2).value, ["v1", "v2"])

    def test_generate_report_no_data(self):
        Robot.objects.all().delete()
        response = self.client.get(self.report_url)

        self.assertEqual(response.status_code, 404)
        response_data = response.json()
        self.assertEqual(response_data['error'], "Нет данных для генерации отчета за последнюю неделю.")